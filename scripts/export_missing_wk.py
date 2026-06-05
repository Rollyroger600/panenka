import sys, re, openpyxl
sys.stdout.reconfigure(encoding='utf-8')

# Parse all dob|country from players.ts
with open('lib/data/players.ts', encoding='utf-8') as f:
    content = f.read()

app_dob_country = set()
for block in re.finditer(r'\{[^}]+\}', content):
    b = block.group()
    country = re.search(r'country: "([^"]+)"', b)
    dob = re.search(r'dob: "([^"]+)"', b)
    if country and dob:
        app_dob_country.add(f'{dob.group(1)}|{country.group(1)}')

# Parse wkOfficialSquads.ts
with open('lib/data/wkOfficialSquads.ts', encoding='utf-8') as f:
    squads_content = f.read()

missing = []
for cb in re.finditer(r"'([^']+)':\s*\[(.*?)\]", squads_content, re.DOTALL):
    country = cb.group(1)
    for entry in re.finditer(r"\{[^}]+\}", cb.group(2)):
        e = entry.group()
        fifa_name = re.search(r"fifaName:\s*'([^']+)'", e)
        dob = re.search(r"dob:\s*'([^']+)'", e)
        position = re.search(r"position:\s*'([^']+)'", e)
        if fifa_name and dob:
            key = f'{dob.group(1)}|{country}'
            if key not in app_dob_country:
                missing.append({
                    'country': country,
                    'fifaName': fifa_name.group(1),
                    'dob': dob.group(1),
                    'position': position.group(1) if position else '?',
                })

missing.sort(key=lambda x: (x['country'], x['dob']))

# Write to Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Ontbrekende WK-spelers'
ws.append(['Land', 'Naam (FIFA)', 'Geboortedatum', 'Positie'])

# Style header
from openpyxl.styles import Font, PatternFill, Alignment
header_fill = PatternFill(start_color='FF6B00', end_color='FF6B00', fill_type='solid')
for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

for p in missing:
    ws.append([p['country'], p['fifaName'], p['dob'], p['position']])

# Auto-width
for col in ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max_len + 4

path = '260605_missing_wk_players.xlsx'
wb.save(path)
print(f'{len(missing)} spelers geëxporteerd naar {path}')
