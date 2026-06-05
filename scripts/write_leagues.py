import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

data = {
    'Abha Club': ('Saudi Pro League', 'Saudi-Arabië', 1),
    'AE Kisia FC': ('Cyprus 3e divisie', 'Cyprus', 3),
    'AEK Larnaca FC': ('Cyprus First Division', 'Cyprus', 1),
    'AEL FC': ('Super League Greece 2', 'Griekenland', 2),
    'AEL Limassol': ('Cyprus First Division', 'Cyprus', 1),
    'Al Bataeh Club': ('UAE Arabian Gulf League', 'VAE', 1),
    'Al Faisaly SC': ('Jordanian Pro League', 'Jordanië', 1),
    'Al Gharafa SC': ('Qatar Stars League', 'Qatar', 1),
    'Al Hussein SC': ('Jordanian Pro League', 'Jordanië', 1),
    'Al Ittihad Kalba SCC': ('UAE First Division League', 'VAE', 2),
    'Al Karma SC': ('Iraqi Premier League', 'Irak', 1),
    'Al Sailiya SC': ('Qatar Stars League', 'Qatar', 1),
    'Al Shamal SC': ('Qatar Stars League', 'Qatar', 1),
    'Al Shorta SC': ('Iraqi Premier League', 'Irak', 1),
    'Al Talaba SC': ('Iraqi Premier League', 'Irak', 1),
    'Al Ula Saudi FC': ('Saudi First Division League', 'Saudi-Arabië', 2),
    'Al Wahda SC': ('UAE Arabian Gulf League', 'VAE', 1),
    'Al Wahdat SC': ('Jordanian Pro League', 'Jordanië', 1),
    "Al Zawra'a SC": ('Iraqi Premier League', 'Irak', 1),
    'Albirex Niigata': ('J1 League', 'Japan', 1),
    'Almere City FC': ('Eredivisie', 'Nederland', 1),
    'Al-Quwa Al-Jawiya': ('Iraqi Premier League', 'Irak', 1),
    'Apollon Limassol': ('Cyprus First Division', 'Cyprus', 1),
    'Aris Limassol FC': ('Cyprus First Division', 'Cyprus', 1),
    'Baniyas Club': ('UAE Arabian Gulf League', 'VAE', 1),
    'CD Cobresal': ('Campeonato Nacional', 'Chili', 1),
    'CD Marathón': ('Liga Nacional de Honduras', 'Honduras', 1),
    'CD Plaza Amador': ('Liga Panameña de Fútbol', 'Panama', 1),
    'CD Universidad De Concepción': ('Campeonato Nacional', 'Chili', 1),
    'Club Africain': ('Ligue Professionnelle 1', 'Tunesië', 1),
    'Club León': ('Liga MX', 'Mexico', 1),
    'Colorado Springs Switchbacks FC': ('USL Championship', 'USA', 2),
    'CR Vasco Da Gama': ('Campeonato Brasileiro Série A', 'Brazilië', 1),
    'Deportivo Saprissa': ('Primera División de Costa Rica', 'Costa Rica', 1),
    'Dibba FC': ('UAE Arabian Gulf League', 'VAE', 1),
    'El Gouna FC': ('Egyptian Premier League', 'Egypte', 1),
    'Espérance De Tunisie': ('Ligue Professionnelle 1', 'Tunesië', 1),
    'Esteghlal Tehran FC': ('Persian Gulf Pro League', 'Iran', 1),
    'Étoile Du Sahel': ('Ligue Professionnelle 1', 'Tunesië', 1),
    'FC Akron Tolyatti': ('Russian Premier League', 'Rusland', 1),
    'FC Astana': ('Kazakhstan Premier League', 'Kazachstan', 1),
    'FC Cosmos Koblenz': ('Regionalliga Südwest', 'Duitsland', 4),
    'FC Den Bosch': ('Eerste Divisie', 'Nederland', 2),
    'FC Dynamo Makhachkala': ('Russian Football National League', 'Rusland', 2),
    'FC Dynamo Moscow': ('Russian Premier League', 'Rusland', 1),
    'FC Hradec Králové': ('Czech First League', 'Tsjechië', 1),
    'FC Noah': ('Armenian Premier League', 'Armenië', 1),
    'FC Pari Nizhny Novgorod': ('Russian Premier League', 'Rusland', 1),
    'FC Rostov': ('Russian Premier League', 'Rusland', 1),
    'FC Slovan Liberec': ('Czech First League', 'Tsjechië', 1),
    'FC Sochaux-Montbéliard': ('Ligue 2', 'Frankrijk', 2),
    'FC Spartak Moscow': ('Russian Premier League', 'Rusland', 1),
    'FC Stade Nyonnais': ('Swiss Promotion League', 'Zwitserland', 3),
    'FC Tatran Prešov': ('Slovak Super Liga', 'Slowakije', 1),
    'FC Tokyo': ('J1 League', 'Japan', 1),
    'FC Vizela': ('Liga Portugal 2', 'Portugal', 2),
    'FK Borac Banja Luka': ('Premijer liga BiH', 'Bosnië', 1),
    'FK Buxoro': ('Uzbek Super League', 'Oezbekistan', 1),
    'FK Dinamo Samarkand': ('Uzbek Super League', 'Oezbekistan', 1),
    "FK Neftchi Farg'ona": ('Uzbek Super League', 'Oezbekistan', 1),
    'Foolad Khuzestan FC': ('Persian Gulf Pro League', 'Iran', 1),
    'GD Chaves': ('Liga Portugal Betclic', 'Portugal', 1),
    'Györi ETO FC': ('OTP Bank Liga', 'Hongarije', 1),
    'Hapoel Kiryat Shmona FC': ('Israeli Premier League', 'Israël', 1),
    'Hearts Of Oak SC': ('Ghana Premier League', 'Ghana', 1),
    'HNK Rijeka': ('HNL - Hrvatska Nogometna Liga', 'Kroatië', 1),
    'Iğdır FK': ('TFF Second League', 'Turkije', 3),
    'JS Kabylie': ('Ligue Professionnelle 1', 'Algerije', 1),
    'Kaizer Chiefs FC': ('DStv Premiership', 'Zuid-Afrika', 1),
    'Kashima Antlers': ('J1 League', 'Japan', 1),
    'Maccabi Haifa FC': ('Israeli Premier League', 'Israël', 1),
    'Malavan Anzali FC': ('Persian Gulf Pro League', 'Iran', 1),
    'Mamelodi Sundowns FC': ('DStv Premiership', 'Zuid-Afrika', 1),
    'Mazatlán FC': ('Liga MX', 'Mexico', 1),
    'Nasaf Qarshi FC': ('Uzbek Super League', 'Oezbekistan', 1),
    'NK Maribor': ('Slovenian PrvaLiga', 'Slovenië', 1),
    'OKMK FK': ('Uzbek Super League', 'Oezbekistan', 1),
    'Orlando Pirates FC': ('DStv Premiership', 'Zuid-Afrika', 1),
    'Pafos FC': ('Cyprus First Division', 'Cyprus', 1),
    'Pakhtakor Tashkent FK': ('Uzbek Super League', 'Oezbekistan', 1),
    'Persepolis FC': ('Persian Gulf Pro League', 'Iran', 1),
    'Persib Bandung': ('Liga 1', 'Indonesië', 1),
    'PFC Ludogorets Razgrad': ('First Professional Football League', 'Bulgarije', 1),
    'PFC Montana': ('Second Professional Football League', 'Bulgarije', 2),
    'PFC Navbahor Namangan': ('Uzbek Super League', 'Oezbekistan', 1),
    'Polokwane City FC': ('DStv Premiership', 'Zuid-Afrika', 1),
    'Port FC': ('Thai League 1', 'Thailand', 1),
    'Pumas UNAM': ('Liga MX', 'Mexico', 1),
    'Puskás Akadémia FC': ('OTP Bank Liga', 'Hongarije', 1),
    'Pyramids FC': ('Egyptian Premier League', 'Egypte', 1),
    'Qatar SC': ('Qatar Stars League', 'Qatar', 1),
    'Raja Casablanca': ('Botola Pro', 'Marokko', 1),
    'Red Bull Bragantino': ('Campeonato Brasileiro Série A', 'Brazilië', 1),
    'RKC Waalwijk': ('Eerste Divisie', 'Nederland', 2),
    'Sanfrecce Hiroshima': ('J1 League', 'Japan', 1),
    'SCU Torreense': ('Liga Portugal 2', 'Portugal', 2),
    'Selangor FC': ('Malaysia Super League', 'Maleisië', 1),
    'Sepahan SC': ('Persian Gulf Pro League', 'Iran', 1),
    'Shabab Al Ahli Club': ('UAE Arabian Gulf League', 'VAE', 1),
    'Siwelele FC': ('Super League of Malawi', 'Malawi', 1),
    'SJK': ('Veikkausliiga', 'Finland', 1),
    'SK Beveren': ('Challenger Pro League', 'België', 2),
    'ŠK Slovan Bratislava': ('Slovak Super Liga', 'Slowakije', 1),
    'Surkhon FK': ('Uzbek Super League', 'Oezbekistan', 1),
    'Terengganu FC': ('Malaysia Super League', 'Maleisië', 1),
    'Tractor Sazi Tabriz FC': ('Persian Gulf Pro League', 'Iran', 1),
    'Turan Tovuz': ('Azerbaijan Premier League', 'Azerbeidzjan', 1),
    'US Monastir': ('Ligue Professionnelle 1', 'Tunesië', 1),
    'USM Alger': ('Ligue Professionnelle 1', 'Algerije', 1),
    'Violette AC': ('Ligue Haïtienne', 'Haïti', 1),
    'VVV Venlo': ('Eerste Divisie', 'Nederland', 2),
    'Zamalek SC': ('Egyptian Premier League', 'Egypte', 1),
    'ZED FC': ('Egyptian Premier League', 'Egypte', 1),
}

wb = openpyxl.load_workbook('docs/reference/draft.xlsx')
ws = wb['Sheet2']

# Update header row: add 'Land' column if needed
ws['C1'] = 'Niveau'

matched = 0
not_found = []

for row in ws.iter_rows(min_row=2):
    club_cell = row[0]
    if club_cell.value:
        club = club_cell.value.strip()
        if club in data:
            comp, land, niveau = data[club]
            row[1].value = comp
            row[2].value = niveau
            matched += 1
        else:
            not_found.append(club)

wb.save('docs/reference/draft.xlsx')
print(f'Bijgewerkt: {matched} clubs')
if not_found:
    print(f'Niet gevonden: {not_found}')
else:
    print('Alle clubs gevonden en bijgewerkt.')
